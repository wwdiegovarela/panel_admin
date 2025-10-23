"""
Diálogo para carga masiva de usuarios desde archivo Excel
"""
from PySide6.QtWidgets import (
    QDialog, QVBoxLayout, QHBoxLayout, QPushButton, QLabel,
    QMessageBox, QTableWidget, QTableWidgetItem, QFileDialog,
    QTextEdit, QGroupBox, QApplication
)
from PySide6.QtCore import Qt
from PySide6.QtGui import QFont
from services.firebase_service import FirebaseService
from services.bigquery_service import BigQueryService
from config.settings import COLOR_PRIMARY, COLOR_SUCCESS, COLOR_ERROR, COLOR_SECONDARY
from ui.loading_dialog import ProgressDialog
from pathlib import Path
import openpyxl
from openpyxl.styles import Font as ExcelFont, PatternFill, Alignment, Border, Side
from datetime import datetime
from typing import List, Dict


class CargaMasivaDialog(QDialog):
    """Diálogo para carga masiva de usuarios desde Excel"""
    
    def __init__(self, parent=None):
        super().__init__(parent)
        self.firebase_service = FirebaseService()
        self.bigquery_service = BigQueryService()
        self.usuarios_validados = []
        self.errores_validacion = []
        
        self.setWindowTitle("Carga Masiva de Usuarios")
        self.setMinimumSize(1000, 700)
        self.init_ui()
    
    def init_ui(self):
        """Inicializar interfaz de usuario"""
        layout = QVBoxLayout(self)
        layout.setContentsMargins(20, 20, 20, 20)
        layout.setSpacing(15)
        
        # Título
        title = QLabel("📊 Carga Masiva de Usuarios")
        title.setStyleSheet(f"font-size: 22px; font-weight: bold; color: {COLOR_PRIMARY};")
        layout.addWidget(title)
        
        # Instrucciones
        instrucciones = QLabel(
            "Descarga la plantilla Excel, complétala con los datos de los usuarios y súbela.\n"
            "El sistema validará los datos antes de crear los usuarios."
        )
        instrucciones.setStyleSheet("color: #666; font-size: 13px;")
        instrucciones.setWordWrap(True)
        layout.addWidget(instrucciones)
        
        # Botones de acción
        botones_layout = QHBoxLayout()
        
        btn_descargar = QPushButton("📥 Descargar Plantilla Excel")
        btn_descargar.clicked.connect(self.descargar_plantilla)
        btn_descargar.setStyleSheet(f"""
            QPushButton {{
                background-color: {COLOR_SECONDARY};
                color: white;
                padding: 12px 20px;
                border-radius: 6px;
                font-weight: bold;
                font-size: 14px;
            }}
            QPushButton:hover {{
                background-color: #d45e0a;
            }}
        """)
        botones_layout.addWidget(btn_descargar)
        
        btn_subir = QPushButton("📤 Subir Archivo Excel")
        btn_subir.clicked.connect(self.subir_archivo)
        btn_subir.setStyleSheet(f"""
            QPushButton {{
                background-color: {COLOR_PRIMARY};
                color: white;
                padding: 12px 20px;
                border-radius: 6px;
                font-weight: bold;
                font-size: 14px;
            }}
            QPushButton:hover {{
                background-color: #025a8a;
            }}
        """)
        botones_layout.addWidget(btn_subir)
        
        botones_layout.addStretch()
        layout.addLayout(botones_layout)
        
        # Vista previa de usuarios
        preview_group = QGroupBox("Vista Previa de Usuarios")
        preview_group.setStyleSheet(f"""
            QGroupBox {{
                font-weight: bold;
                border: 2px solid {COLOR_PRIMARY};
                border-radius: 6px;
                margin-top: 10px;
                padding-top: 10px;
            }}
            QGroupBox::title {{
                subcontrol-origin: margin;
                left: 10px;
                padding: 0 5px;
            }}
        """)
        preview_layout = QVBoxLayout()
        
        self.tabla_preview = QTableWidget()
        self.tabla_preview.setColumnCount(7)
        self.tabla_preview.setHorizontalHeaderLabels([
            "Email", "Nombre", "Cargo", "Teléfono", "Rol", "Es Contacto", "# Instalaciones"
        ])
        self.tabla_preview.horizontalHeader().setStretchLastSection(True)
        self.tabla_preview.setAlternatingRowColors(True)
        preview_layout.addWidget(self.tabla_preview)
        
        preview_group.setLayout(preview_layout)
        layout.addWidget(preview_group)
        
        # Log de errores
        errores_group = QGroupBox("Errores de Validación")
        errores_group.setStyleSheet(f"""
            QGroupBox {{
                font-weight: bold;
                border: 2px solid {COLOR_ERROR};
                border-radius: 6px;
                margin-top: 10px;
                padding-top: 10px;
            }}
            QGroupBox::title {{
                subcontrol-origin: margin;
                left: 10px;
                padding: 0 5px;
            }}
        """)
        errores_layout = QVBoxLayout()
        
        self.log_errores = QTextEdit()
        self.log_errores.setReadOnly(True)
        self.log_errores.setMaximumHeight(150)
        self.log_errores.setStyleSheet("background-color: #fff3cd; color: #856404;")
        errores_layout.addWidget(self.log_errores)
        
        errores_group.setLayout(errores_layout)
        layout.addWidget(errores_group)
        
        # Botones de acción final
        botones_finales = QHBoxLayout()
        botones_finales.addStretch()
        
        self.btn_crear = QPushButton("✅ Crear Usuarios")
        self.btn_crear.clicked.connect(self.crear_usuarios)
        self.btn_crear.setEnabled(False)
        self.btn_crear.setStyleSheet(f"""
            QPushButton {{
                background-color: {COLOR_SUCCESS};
                color: white;
                padding: 12px 30px;
                border-radius: 6px;
                font-weight: bold;
                font-size: 14px;
            }}
            QPushButton:hover {{
                background-color: #45a049;
            }}
            QPushButton:disabled {{
                background-color: #cccccc;
                color: #666666;
            }}
        """)
        botones_finales.addWidget(self.btn_crear)
        
        btn_cancelar = QPushButton("❌ Cancelar")
        btn_cancelar.clicked.connect(self.reject)
        btn_cancelar.setStyleSheet("""
            QPushButton {
                background-color: #dc3545;
                color: white;
                padding: 12px 30px;
                border-radius: 6px;
                font-weight: bold;
                font-size: 14px;
            }
            QPushButton:hover {
                background-color: #c82333;
            }
        """)
        botones_finales.addWidget(btn_cancelar)
        
        layout.addLayout(botones_finales)
    
    def descargar_plantilla(self):
        """Generar y descargar plantilla Excel"""
        try:
            # Pedir al usuario dónde guardar
            ruta_guardar, _ = QFileDialog.getSaveFileName(
                self,
                "Guardar Plantilla Excel",
                str(Path.home() / f"plantilla_usuarios_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"),
                "Excel Files (*.xlsx)"
            )
            
            if not ruta_guardar:
                return
            
            # Intentar usar XlsxWriter (mejor compatibilidad con listas desplegables)
            try:
                import xlsxwriter  # type: ignore
                wb = xlsxwriter.Workbook(ruta_guardar)

                # ===== HOJA 1: USUARIOS =====
                ws_usuarios = wb.add_worksheet("Usuarios")

                header_fmt = wb.add_format({
                    'bold': True, 'font_color': 'white', 'bg_color': '#0275AA',
                    'align': 'center', 'valign': 'vcenter'
                })
                border_fmt = wb.add_format({'border': 1})

                headers = [
                    "Email *", "Nombre Completo *", "Cargo", "Teléfono",
                    "Contraseña (min 6, opcional)", "Rol *", "Es Contacto (SI/NO) *", "Instalaciones (separadas por coma) *"
                ]
                for col, header in enumerate(headers):
                    ws_usuarios.write(0, col, header, header_fmt)
                ws_usuarios.set_column('A:A', 30)
                ws_usuarios.set_column('B:B', 25)
                ws_usuarios.set_column('C:C', 20)
                ws_usuarios.set_column('D:D', 15)
                ws_usuarios.set_column('E:E', 18)  # Contraseña
                ws_usuarios.set_column('F:F', 15)  # Rol
                ws_usuarios.set_column('G:G', 20)  # Es contacto
                ws_usuarios.set_column('H:H', 50)  # Instalaciones

                # Ejemplos
                ejemplos = [
                    ["juan.perez@empresa.cl", "Juan Pérez", "Supervisor", "+56912345678", "Temp1234", "CLIENTE", "NO", "PLANTA_A,PLANTA_B,OFICINA_CENTRAL"],
                    ["maria.lopez@wfsa.cl", "María López", "Administradora", "+56987654321", "", "ADMIN_WFSA", "NO", ""]
                ]
                for r, ejemplo in enumerate(ejemplos, start=1):
                    for c, v in enumerate(ejemplo):
                        ws_usuarios.write(r, c, v, border_fmt)

                # ===== HOJA 2: ROLES =====
                ws_roles = wb.add_worksheet("Roles")
                ws_roles.write_row(0, 0, ["rol_id", "nombre", "descripcion"], header_fmt)
                roles = self.bigquery_service.get_roles()
                for idx, rol in enumerate(roles, start=1):
                    ws_roles.write_row(idx, 0, [
                        rol.get('rol_id', ''),
                        rol.get('nombre_rol', ''),
                        rol.get('descripcion', '')
                    ], border_fmt)
                ws_roles.set_column('A:A', 20)
                ws_roles.set_column('B:B', 35)
                ws_roles.set_column('C:C', 60)

                # Validación: Rol (F2:F5000) usando lista oculta en la misma hoja (columna J)
                if roles:
                    roles_ids = [r.get('rol_id', '') for r in roles]
                    if roles_ids:
                        ws_usuarios.write_column('J2', roles_ids)
                        ws_usuarios.set_column('J:J', None, None, {'hidden': True})
                        last_row_roles = len(roles_ids) + 1
                        source_range = f"=$J$2:$J${last_row_roles}"
                        ws_usuarios.data_validation('F2:F5000', {
                            'validate': 'list',
                            'source': source_range,
                            'error_title': 'Rol inválido',
                            'error_message': 'Selecciona un rol de la lista',
                        })

                # Validación: Es Contacto (G2:G5000)
                ws_usuarios.data_validation('G2:G5000', {
                    'validate': 'list',
                    'source': ['SI', 'NO'],
                    'error_title': 'Valor inválido',
                    'error_message': 'Elige SI o NO'
                })

                # ===== HOJA 3: INSTALACIONES =====
                ws_inst = wb.add_worksheet("Instalaciones Disponibles")
                ws_inst.write_row(0, 0, ["Instalación", "Cliente", "Zona", "Comuna"], header_fmt)
                instalaciones = self.bigquery_service.get_instalaciones_con_zonas()
                for idx, inst in enumerate(instalaciones, start=1):
                    ws_inst.write_row(idx, 0, [
                        inst.get('instalacion_rol', ''),
                        inst.get('cliente_rol', ''),
                        inst.get('zona', ''),
                        inst.get('comuna', '')
                    ], border_fmt)
                ws_inst.set_column('A:A', 30)
                ws_inst.set_column('B:B', 25)
                ws_inst.set_column('C:C', 20)
                ws_inst.set_column('D:D', 20)

                # ===== HOJA 4: INSTRUCCIONES =====
                ws_ins = wb.add_worksheet("Instrucciones")
                instrucciones_texto = [
                    "INSTRUCCIONES DE USO - CARGA MASIVA DE USUARIOS",
                    "",
                    "1. COMPLETAR LA HOJA 'Usuarios'",
                    "   - Email (obligatorio)",
                    "   - Nombre Completo (obligatorio)",
                    "   - Cargo (opcional)",
                    "   - Teléfono",
                    "   - Rol: usar la lista desplegable (columna E)",
                    "   - Es Contacto: usar SI/NO (columna F)",
                    "   - Instalaciones: separadas por coma",
                    "",
                    "2. VALIDACIONES IMPORTANTES",
                    "   - CLIENTE no puede ser 'Es Contacto' = SI",
                    "   - No @wfsa.cl => solo rol CLIENTE",
                    "   - Emails únicos",
                    "",
                    "3. ROLES DISPONIBLES",
                    "   - Ver hoja 'Roles'",
                    "",
                    "4. INSTALACIONES",
                    "   - Ver hoja 'Instalaciones Disponibles' (incluye Cliente y Zona)",
                ]
                for i, t in enumerate(instrucciones_texto):
                    ws_ins.write(i, 0, t)

                wb.close()

                # Mensaje de éxito y salir sin ejecutar el fallback
                QMessageBox.information(
                    self,
                    "Plantilla Descargada",
                    f"La plantilla se descargó correctamente en:\n\n{ruta_guardar}\n\n"
                    "Completa los datos y súbela usando el botón 'Subir Archivo Excel'."
                )
                return

            except Exception:
                # Fallback a openpyxl
                # Crear workbook
                wb = openpyxl.Workbook()
            
            # ===== HOJA 1: USUARIOS =====
            ws_usuarios = wb.active
            ws_usuarios.title = "Usuarios"
            
            # Estilos
            header_fill = PatternFill(start_color="0275AA", end_color="0275AA", fill_type="solid")
            header_font = ExcelFont(color="FFFFFF", bold=True, size=12)
            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # Encabezados
            headers = [
                "Email *", "Nombre Completo *", "Cargo", "Teléfono", 
                "Contraseña (min 6, opcional)", "Rol *", "Es Contacto (SI/NO) *", "Instalaciones (separadas por coma) *"
            ]
            
            for col, header in enumerate(headers, start=1):
                cell = ws_usuarios.cell(row=1, column=col)
                cell.value = header
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = border
            
            # Ajustar ancho de columnas
            ws_usuarios.column_dimensions['A'].width = 30
            ws_usuarios.column_dimensions['B'].width = 25
            ws_usuarios.column_dimensions['C'].width = 20
            ws_usuarios.column_dimensions['D'].width = 15
            ws_usuarios.column_dimensions['E'].width = 18  # Contraseña
            ws_usuarios.column_dimensions['F'].width = 15  # Rol
            ws_usuarios.column_dimensions['G'].width = 20  # Es contacto
            ws_usuarios.column_dimensions['H'].width = 50  # Instalaciones
            
            # Ejemplos (2 filas de ejemplo)
            ejemplos = [
                ["juan.perez@empresa.cl", "Juan Pérez", "Supervisor", "+56912345678", "Temp1234", "CLIENTE", "NO", "PLANTA_A,PLANTA_B,OFICINA_CENTRAL"],
                ["maria.lopez@wfsa.cl", "María López", "Administradora", "+56987654321", "", "ADMIN_WFSA", "NO", ""]
            ]
            
            for row_idx, ejemplo in enumerate(ejemplos, start=2):
                for col_idx, valor in enumerate(ejemplo, start=1):
                    cell = ws_usuarios.cell(row=row_idx, column=col_idx)
                    cell.value = valor
                    cell.border = border
                    cell.alignment = Alignment(vertical='center')
            
            # ===== HOJA 2: ROLES (referencia informativa) =====
            ws_roles = wb.create_sheet("Roles")
            ws_roles.append(["rol_id", "nombre", "descripcion"])        
            roles = self.bigquery_service.get_roles()
            for rol in roles:
                ws_roles.append([
                    rol.get('rol_id', ''),
                    rol.get('nombre_rol', ''),
                    rol.get('descripcion', '')
                ])
            ws_roles.column_dimensions['A'].width = 20
            ws_roles.column_dimensions['B'].width = 35
            ws_roles.column_dimensions['C'].width = 60

            # Validación de datos para columna Rol en hoja Usuarios (lista desplegable)
            # Nota: Para máxima compatibilidad con Excel, usamos un rango en la misma hoja ('Usuarios')
            # que se oculta, evitando restricciones de listas desde otras hojas.
            try:
                from openpyxl.worksheet.datavalidation import DataValidation
                if roles:
                    # Escribir lista de roles en columna oculta J (desde J2)
                    start_row = 2
                    for idx, rol in enumerate(roles, start=start_row):
                        ws_usuarios.cell(row=idx, column=10).value = rol.get('rol_id', '')  # Columna J = 10
                    # Ocultar columna J
                    ws_usuarios.column_dimensions['J'].hidden = True
                    # Definir validación apuntando al rango en la misma hoja
                    last_row_roles = len(roles) + start_row - 1
                    formula_range = f"$J$2:$J${last_row_roles}"
                    dv_roles = DataValidation(type="list", formula1=f"={formula_range}", allow_blank=False, showDropDown=True)
                    dv_roles.error = "Rol inválido. Selecciona un valor de la lista."
                    dv_roles.prompt = "Selecciona un rol válido"
                    ws_usuarios.add_data_validation(dv_roles)
                    dv_roles.add("F2:F5000")
            except Exception:
                pass

            # Validación de datos para 'Es Contacto' (SI/NO)
            try:
                from openpyxl.worksheet.datavalidation import DataValidation
                dv_contacto = DataValidation(type="list", formula1='"SI,NO"', allow_blank=False, showDropDown=True)
                dv_contacto.error = "Valor inválido. Usa SI o NO"
                dv_contacto.prompt = "Elige SI o NO"
                ws_usuarios.add_data_validation(dv_contacto)
                dv_contacto.add("G2:G5000")
            except Exception:
                pass

            # ===== HOJA 3: INSTALACIONES =====
            ws_instalaciones = wb.create_sheet("Instalaciones Disponibles")
            
            # Obtener instalaciones con zonas
            instalaciones = self.bigquery_service.get_instalaciones_con_zonas()
            
            # Encabezados
            headers_inst = ["Instalación", "Cliente", "Zona", "Comuna"]
            for col, header in enumerate(headers_inst, start=1):
                cell = ws_instalaciones.cell(row=1, column=col)
                cell.value = header
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = border
            
            # Datos
            for row_idx, inst in enumerate(instalaciones, start=2):
                datos = [
                    inst.get('instalacion_rol', ''),
                    inst.get('cliente_rol', ''),
                    inst.get('zona', ''),
                    inst.get('comuna', '')
                ]
                for col_idx, valor in enumerate(datos, start=1):
                    cell = ws_instalaciones.cell(row=row_idx, column=col_idx)
                    cell.value = valor
                    cell.border = border
            
            # Ajustar ancho de columnas
            ws_instalaciones.column_dimensions['A'].width = 30
            ws_instalaciones.column_dimensions['B'].width = 25
            ws_instalaciones.column_dimensions['C'].width = 20
            ws_instalaciones.column_dimensions['D'].width = 20
            
            # ===== HOJA 4: INSTRUCCIONES =====
            ws_instrucciones = wb.create_sheet("Instrucciones")
            
            instrucciones_texto = [
                ("INSTRUCCIONES DE USO - CARGA MASIVA DE USUARIOS", True, 16),
                ("", False, 11),
                ("1. COMPLETAR LA HOJA 'Usuarios'", True, 13),
                ("   - Email: Correo electrónico del usuario (único, obligatorio)", False, 11),
                ("   - Nombre Completo: Nombre y apellido del usuario", False, 11),
                ("   - Cargo: Cargo o puesto del usuario (opcional)", False, 11),
                ("   - Teléfono: Número de teléfono con formato +56912345678", False, 11),
                ("   - Rol: Código del rol (ver roles disponibles más abajo)", False, 11),
                ("   - Es Contacto: 'SI' o 'NO' (si será contacto de WhatsApp)", False, 11),
                ("   - Instalaciones: Códigos de instalación separados por coma", False, 11),
                ("", False, 11),
                ("2. VALIDACIONES IMPORTANTES", True, 13),
                ("   ⚠️ Si el rol es 'CLIENTE' NO puede tener 'Es Contacto' = 'SI'", False, 11),
                ("   ⚠️ Si el email NO es '@wfsa.cl' SOLO puede tener rol 'CLIENTE'", False, 11),
                ("   ⚠️ Los emails deben ser únicos (no duplicados)", False, 11),
                ("   ⚠️ Las instalaciones deben existir (ver hoja 'Instalaciones Disponibles')", False, 11),
                ("", False, 11),
                ("3. ROLES DISPONIBLES", True, 13),
            ]
            
            # Obtener roles
            roles = self.bigquery_service.get_roles()
            for rol in roles:
                instrucciones_texto.append((
                    f"   • {rol.get('rol_id')}: {rol.get('nombre_rol')} - {rol.get('descripcion', '')}",
                    False,
                    11
                ))
            
            instrucciones_texto.extend([
                ("", False, 11),
                ("4. CONSULTAR INSTALACIONES", True, 13),
                ("   - Ver hoja 'Instalaciones Disponibles' para obtener los códigos exactos", False, 11),
                ("   - Copiar y pegar los códigos separados por coma (sin espacios)", False, 11),
                ("   - Ejemplo: PLANTA_A,PLANTA_B,OFICINA_CENTRAL", False, 11),
                ("", False, 11),
                ("5. SUBIR EL ARCHIVO", True, 13),
                ("   - Guardar el archivo después de completarlo", False, 11),
                ("   - En el Panel Admin, hacer clic en 'Subir Archivo Excel'", False, 11),
                ("   - El sistema validará los datos automáticamente", False, 11),
                ("   - Revisar errores si los hay y corregir", False, 11),
                ("   - Si todo está correcto, hacer clic en 'Crear Usuarios'", False, 11),
                ("", False, 11),
                ("SOPORTE: Si tienes dudas, contacta al administrador del sistema.", True, 12),
            ])
            
            # Escribir instrucciones
            for row_idx, (texto, bold, size) in enumerate(instrucciones_texto, start=1):
                cell = ws_instrucciones.cell(row=row_idx, column=1)
                cell.value = texto
                cell.font = ExcelFont(bold=bold, size=size)
                cell.alignment = Alignment(vertical='center', wrap_text=True)
            
            ws_instrucciones.column_dimensions['A'].width = 100
            
            # Guardar archivo
            wb.save(ruta_guardar)
            
            QMessageBox.information(
                self,
                "Plantilla Descargada",
                f"La plantilla se descargó correctamente en:\n\n{ruta_guardar}\n\n"
                "Completa los datos y súbela usando el botón 'Subir Archivo Excel'."
            )
            
        except Exception as e:
            QMessageBox.critical(
                self,
                "Error",
                f"No se pudo generar la plantilla:\n{str(e)}"
            )
    
    def subir_archivo(self):
        """Subir y validar archivo Excel"""
        try:
            # Seleccionar archivo
            ruta_archivo, _ = QFileDialog.getOpenFileName(
                self,
                "Seleccionar Archivo Excel",
                str(Path.home()),
                "Excel Files (*.xlsx *.xls)"
            )
            
            if not ruta_archivo:
                return
            
            # Rechazar .xls (no soportado por openpyxl)
            if ruta_archivo.lower().endswith('.xls'):
                QMessageBox.warning(
                    self,
                    "Formato no soportado",
                    "El archivo debe estar en formato .xlsx.\n\n"
                    "Abre el archivo en Excel y guárdalo como 'Libro de Excel (*.xlsx)'."
                )
                return
            
            # Leer archivo
            self.log_errores.clear()
            self.log_errores.append("📂 Leyendo archivo...")
            QApplication.processEvents()
            
            wb = openpyxl.load_workbook(ruta_archivo, data_only=True)
            # Buscar hoja 'Usuarios' case-insensitive
            hoja_usuarios = None
            for name in wb.sheetnames:
                if name.strip().lower() == 'usuarios':
                    hoja_usuarios = name
                    break
            if not hoja_usuarios:
                # Fallback: primera hoja
                hoja_usuarios = wb.sheetnames[0]
            ws = wb[hoja_usuarios]
            
            # Leer datos (detección por encabezados, tolerante a orden/ediciones)
            usuarios_raw = []
            ejemplo_emails = {"juan.perez@empresa.cl", "maria.lopez@wfsa.cl"}

            # Buscar fila de encabezados en primeras 10 filas
            header_row_idx = None
            header_map = {}
            expected_keys = {
                'email': 'email',
                'nombre': 'nombre',
                'cargo': 'cargo',
                'telefono': 'telefono',
                'teléfono': 'telefono',
                'contraseña': 'password',
                'contrasena': 'password',
                'rol': 'rol',
                'es contacto': 'es_contacto',
                'instalaciones': 'instalaciones',
            }
            for i, row in enumerate(ws.iter_rows(min_row=1, max_row=10, values_only=True), start=1):
                if not row:
                    continue
                labels = [str(c).strip().lower() if isinstance(c, str) else '' for c in row]
                if any('email' in lbl for lbl in labels):
                    header_row_idx = i
                    for idx, lbl in enumerate(labels):
                        for key, mapped in expected_keys.items():
                            if lbl.startswith(key):
                                header_map[mapped] = idx
                    break

            # Si no se detecta header, asumir A..G y empezar en fila 2
            if header_row_idx is None or 'email' not in header_map:
                header_row_idx = 1
                header_map = {
                    'email': 0,
                    'nombre': 1,
                    'cargo': 2,
                    'telefono': 3,
                    'password': 4,
                    'rol': 5,
                    'es_contacto': 6,
                    'instalaciones': 7,
                }

            # Recorrer datos desde la fila siguiente al encabezado
            for row_idx, row in enumerate(ws.iter_rows(min_row=header_row_idx + 1, values_only=True), start=header_row_idx + 1):
                if not row or not any(row):
                    continue
                def get_col(key):
                    idx = header_map.get(key)
                    return (str(row[idx]).strip() if (idx is not None and idx < len(row) and row[idx] is not None) else '')

                email = get_col('email')
                if not email:
                    continue
                # Saltar ejemplos
                if email in ejemplo_emails:
                    continue

                usuario = {
                    'fila': row_idx,
                    'email': email,
                    'nombre': get_col('nombre'),
                    'cargo': get_col('cargo'),
                    'telefono': get_col('telefono'),
                    'password': get_col('password'),
                    'rol': get_col('rol'),
                    'es_contacto': get_col('es_contacto').upper() or 'NO',
                    'instalaciones': get_col('instalaciones')
                }
                usuarios_raw.append(usuario)
            
            if not usuarios_raw:
                QMessageBox.warning(
                    self,
                    "Archivo Vacío",
                    "No se encontraron usuarios en el archivo.\n\n"
                    "Asegúrate de completar la hoja 'Usuarios' con los datos."
                )
                return
            
            self.log_errores.append(f"✅ Se encontraron {len(usuarios_raw)} usuarios en el archivo")
            self.log_errores.append("\n🔍 Validando datos...")
            QApplication.processEvents()
            
            # Validar usuarios
            self.usuarios_validados, self.errores_validacion = self.validar_usuarios(usuarios_raw)
            
            # Mostrar resultados
            if self.errores_validacion:
                self.log_errores.append(f"\n❌ Se encontraron {len(self.errores_validacion)} errores:\n")
                for error in self.errores_validacion:
                    self.log_errores.append(f"   • {error}")
                
                self.btn_crear.setEnabled(False)
            else:
                self.log_errores.append(f"\n✅ Todos los datos son válidos!")
                self.btn_crear.setEnabled(True)
            
            # Mostrar vista previa
            self.mostrar_preview()
            
        except Exception as e:
            QMessageBox.critical(
                self,
                "Error",
                f"No se pudo leer el archivo:\n{str(e)}"
            )
    
    def validar_usuarios(self, usuarios_raw: List[Dict]) -> tuple[List[Dict], List[str]]:
        """
        Validar usuarios del archivo Excel
        
        Returns:
            (usuarios_validos, errores)
        """
        errores = []
        usuarios_validos = []
        
        # Obtener datos de validación
        roles_validos = {r['rol_id']: r for r in self.bigquery_service.get_roles()}
        instalaciones_existentes = {i['instalacion_rol']: i for i in self.bigquery_service.get_instalaciones()}
        usuarios_existentes = {u['email_login'] for u in self.bigquery_service.get_usuarios_con_roles()}
        
        emails_en_archivo = set()
        
        for idx, usuario in enumerate(usuarios_raw, start=1):
            fila = usuario['fila']
            errores_usuario = []
            
            # Validar email
            if not usuario['email']:
                errores_usuario.append(f"Fila {fila}: Email es obligatorio")
            elif usuario['email'] in emails_en_archivo:
                errores_usuario.append(f"Fila {fila}: Email '{usuario['email']}' está duplicado en el archivo")
            elif usuario['email'] in usuarios_existentes:
                errores_usuario.append(f"Fila {fila}: Email '{usuario['email']}' ya existe en el sistema")
            else:
                emails_en_archivo.add(usuario['email'])
            
            # Validar nombre
            if not usuario['nombre']:
                errores_usuario.append(f"Fila {fila}: Nombre es obligatorio")
            
            # Validar rol
            if not usuario['rol']:
                errores_usuario.append(f"Fila {fila}: Rol es obligatorio")
            elif usuario['rol'] not in roles_validos:
                errores_usuario.append(f"Fila {fila}: Rol '{usuario['rol']}' no es válido")
            
            # Validar es_contacto
            if usuario['es_contacto'] not in ['SI', 'NO']:
                errores_usuario.append(f"Fila {fila}: 'Es Contacto' debe ser 'SI' o 'NO'")
            
            # Validar instalaciones
            if not usuario['instalaciones']:
                errores_usuario.append(f"Fila {fila}: Debe especificar al menos una instalación")
            else:
                instalaciones_list = [i.strip() for i in usuario['instalaciones'].split(',')]
                for instalacion in instalaciones_list:
                    if instalacion not in instalaciones_existentes:
                        errores_usuario.append(f"Fila {fila}: Instalación '{instalacion}' no existe")
            
            # VALIDACIÓN ESPECIAL 1: CLIENTE no puede ser contacto
            if usuario['rol'] == 'CLIENTE' and usuario['es_contacto'] == 'SI':
                errores_usuario.append(
                    f"Fila {fila}: Un usuario con rol CLIENTE NO puede tener 'Es Contacto' = SI"
                )
            
            # VALIDACIÓN ESPECIAL 2: Solo @wfsa.cl puede tener roles que no sean CLIENTE
            if usuario['email'] and not usuario['email'].endswith('@wfsa.cl'):
                if usuario['rol'] and usuario['rol'] != 'CLIENTE':
                    errores_usuario.append(
                        f"Fila {fila}: Email '{usuario['email']}' no es @wfsa.cl, "
                        f"solo puede tener rol CLIENTE (tiene '{usuario['rol']}')"
                    )
            
            # Si hay errores, agregarlos y continuar
            if errores_usuario:
                errores.extend(errores_usuario)
            else:
                # Procesar instalaciones y extraer clientes
                instalaciones_list = [i.strip() for i in usuario['instalaciones'].split(',')]
                instalaciones_con_cliente = {}
                clientes = set()
                
                for instalacion in instalaciones_list:
                    if instalacion in instalaciones_existentes:
                        cliente = instalaciones_existentes[instalacion].get('cliente_rol', '')
                        instalaciones_con_cliente[instalacion] = cliente
                        if cliente:
                            clientes.add(cliente)
                
                usuario['instalaciones_con_cliente'] = instalaciones_con_cliente
                usuario['cliente_rol'] = ','.join(sorted(clientes))
                usuarios_validos.append(usuario)
        
        return usuarios_validos, errores
    
    def mostrar_preview(self):
        """Mostrar vista previa de usuarios validados"""
        self.tabla_preview.setRowCount(len(self.usuarios_validados))
        
        for row, usuario in enumerate(self.usuarios_validados):
            self.tabla_preview.setItem(row, 0, QTableWidgetItem(usuario['email']))
            self.tabla_preview.setItem(row, 1, QTableWidgetItem(usuario['nombre']))
            self.tabla_preview.setItem(row, 2, QTableWidgetItem(usuario['cargo']))
            self.tabla_preview.setItem(row, 3, QTableWidgetItem(usuario['telefono']))
            self.tabla_preview.setItem(row, 4, QTableWidgetItem(usuario['rol']))
            self.tabla_preview.setItem(row, 5, QTableWidgetItem(usuario['es_contacto']))
            self.tabla_preview.setItem(row, 6, QTableWidgetItem(str(len(usuario['instalaciones_con_cliente']))))
    
    def crear_usuarios(self):
        """Crear usuarios en batch"""
        if not self.usuarios_validados:
            return
        
        respuesta = QMessageBox.question(
            self,
            "Confirmar Creación",
            f"¿Estás seguro de crear {len(self.usuarios_validados)} usuarios?\n\n"
            "Esta acción no se puede deshacer.",
            QMessageBox.Yes | QMessageBox.No
        )
        
        if respuesta != QMessageBox.Yes:
            return
        
        # Crear barra de progreso
        total_pasos = len(self.usuarios_validados) * 4  # Firebase + BigQuery + Instalaciones + Contacto (si aplica)
        progress = ProgressDialog(self, "Creando Usuarios", total_pasos)
        progress.show()
        QApplication.processEvents()
        
        usuarios_creados = 0
        usuarios_fallidos = []
        paso_actual = 0
        
        for idx, usuario in enumerate(self.usuarios_validados, start=1):
            try:
                # Paso 1: Crear en Firebase
                paso_actual += 1
                progress.update_progress(
                    paso_actual,
                    f"[{idx}/{len(self.usuarios_validados)}] Creando en Firebase: {usuario['email']}"
                )
                
                password_eff = usuario.get('password') or "TempPassword123!"
                firebase_result = self.firebase_service.create_user(
                    email=usuario['email'],
                    password=password_eff,
                    display_name=usuario['nombre']
                )
                
                if not firebase_result['success']:
                    raise Exception(f"Firebase: {firebase_result.get('error')}")
                
                # Paso 2: Crear en BigQuery
                paso_actual += 1
                progress.update_progress(
                    paso_actual,
                    f"[{idx}/{len(self.usuarios_validados)}] Guardando en BigQuery: {usuario['email']}"
                )
                
                bq_result = self.bigquery_service.create_usuario(
                    email=usuario['email'],
                    firebase_uid=firebase_result['uid'],
                    cliente_rol=usuario['cliente_rol'],
                    nombre_completo=usuario['nombre'],
                    rol_id=usuario['rol'],
                    cargo=usuario['cargo'],
                    telefono=usuario['telefono'],
                    ver_todas_instalaciones=False
                )
                
                if not bq_result['success']:
                    raise Exception(f"BigQuery: {bq_result.get('error')}")
                
                # Paso 3: Asignar instalaciones
                paso_actual += 1
                progress.update_progress(
                    paso_actual,
                    f"[{idx}/{len(self.usuarios_validados)}] Asignando instalaciones: {usuario['email']}"
                )
                
                inst_result = self.bigquery_service.asignar_instalaciones_multi_cliente(
                    usuario['email'],
                    usuario['instalaciones_con_cliente']
                )
                
                if not inst_result['success']:
                    raise Exception(f"Instalaciones: {inst_result.get('error')}")
                
                # Paso 4: Crear contacto si es necesario
                paso_actual += 1
                if usuario['es_contacto'] == 'SI':
                    progress.update_progress(
                        paso_actual,
                        f"[{idx}/{len(self.usuarios_validados)}] Creando contacto: {usuario['email']}"
                    )
                    
                    contacto_result = self.bigquery_service.create_contacto(
                        nombre=usuario['nombre'],
                        telefono=usuario['telefono'],
                        cargo=usuario['cargo'],
                        email=usuario['email'],
                        es_usuario_app=True
                    )
                    
                    if contacto_result['success']:
                        # Sincronizar instalaciones
                        sync_result = self.bigquery_service.sincronizar_instalaciones_contacto(
                            usuario['email'],
                            usuario['instalaciones_con_cliente']
                        )
                else:
                    progress.update_progress(
                        paso_actual,
                        f"[{idx}/{len(self.usuarios_validados)}] Omitiendo contacto para: {usuario['email']}"
                    )
                
                usuarios_creados += 1
                
            except Exception as e:
                usuarios_fallidos.append(f"{usuario['email']}: {str(e)}")
        
        progress.close()
        
        # Mostrar resumen
        if usuarios_fallidos:
            mensaje = (
                f"✅ {usuarios_creados} usuarios creados correctamente\n"
                f"❌ {len(usuarios_fallidos)} usuarios fallaron:\n\n"
            )
            mensaje += "\n".join(f"• {error}" for error in usuarios_fallidos[:10])
            if len(usuarios_fallidos) > 10:
                mensaje += f"\n... y {len(usuarios_fallidos) - 10} más"
            
            QMessageBox.warning(self, "Creación Parcial", mensaje)
        else:
            QMessageBox.information(
                self,
                "Éxito",
                f"✅ Se crearon correctamente {usuarios_creados} usuarios!\n\n"
                "Los usuarios recibirán un correo para establecer su contraseña."
            )
        
        self.accept()

